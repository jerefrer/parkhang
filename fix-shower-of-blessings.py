#!/usr/bin/env python3
"""
Fix shower-of-blessings.xlsx to match the expected format.

This script fixes two issues:
1. Translation title was in column 4 (should be in columns 2 & 3)
2. Empty row 6 between title rows and header row

Run this script if you need to re-apply the fix:
    python3 fix-shower-of-blessings.py
"""

import openpyxl
from openpyxl import load_workbook
import sys

def fix_shower_of_blessings():
    try:
        # Load the file
        wb = load_workbook('pechas/shower-of-blessings.xlsx')
        ws = wb.active
        
        print("Fixing shower-of-blessings.xlsx...")
        print(f"Before: Row 4 = {[ws.cell(4, i).value for i in range(1, 6)]}")
        
        # Fix Row 4: Move French title from column 4 to columns 2 & 3
        french_title = ws.cell(4, 4).value
        if french_title:
            ws.cell(4, 2).value = french_title  # English column
            ws.cell(4, 3).value = french_title  # French column
            ws.cell(4, 4).value = None  # Clear old position
        
        # Check if row 6 is empty and delete it
        row6_values = [ws.cell(6, i).value for i in range(1, 5)]
        if not any(row6_values):
            print("Deleting empty row 6...")
            ws.delete_rows(6, 1)
        
        print(f"After:  Row 4 = {[ws.cell(4, i).value for i in range(1, 6)]}")
        print(f"        Row 6 = {[ws.cell(6, i).value for i in range(1, 6)]}")
        
        # Save the fixed file
        wb.save('pechas/shower-of-blessings.xlsx')
        print("\n✅ File fixed successfully!")
        return 0
        
    except FileNotFoundError:
        print("❌ Error: pechas/shower-of-blessings.xlsx not found")
        return 1
    except Exception as e:
        print(f"❌ Error: {e}")
        return 1

if __name__ == "__main__":
    sys.exit(fix_shower_of_blessings())
